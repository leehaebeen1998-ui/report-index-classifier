from __future__ import annotations

import base64
import csv
import io
import os
import subprocess
from pathlib import Path
from typing import Any


def read_report_rows(path: str | Path, *, sheet_name: str | None = None) -> list[dict[str, Any]]:
    report_path = Path(path)
    suffix = report_path.suffix.casefold()

    if suffix == ".csv":
        return _read_delimited(report_path, delimiter=",")
    if suffix == ".tsv":
        return _read_delimited(report_path, delimiter="\t")
    if suffix == ".xlsx":
        return _read_xlsx(report_path, sheet_name=sheet_name)

    raise ValueError(f"Unsupported report file type: {report_path.suffix}")


def write_csv_rows(
    path: str | Path,
    rows: list[dict[str, Any]],
    *,
    fieldnames: list[str],
    encoding: str = "utf-8-sig",
) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    file = io.StringIO(newline="")
    writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    write_text(output_path, file.getvalue(), encoding=encoding)


def write_text(path: str | Path, content: str, *, encoding: str = "utf-8") -> None:
    output_path = Path(path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with output_path.open("w", encoding=encoding, newline="") as file:
            file.write(content)
    except OSError:
        _write_text_with_powershell(output_path, content, encoding=encoding)


def _write_text_with_powershell(path: Path, content: str, *, encoding: str) -> None:
    command = (
        "$target = $env:INDEX_CLASSIFIER_TARGET_PATH; "
        "$directory = [System.IO.Path]::GetDirectoryName($target); "
        "if ($directory) { [System.IO.Directory]::CreateDirectory($directory) | Out-Null }; "
        "$base64 = [Console]::In.ReadToEnd(); "
        "$bytes = [Convert]::FromBase64String($base64); "
        "[System.IO.File]::WriteAllBytes($target, $bytes)"
    )
    encoded = base64.b64encode(content.encode(encoding)).decode("ascii")
    env = os.environ.copy()
    env["INDEX_CLASSIFIER_TARGET_PATH"] = os.fspath(path)
    result = subprocess.run(
        ["powershell.exe", "-NoProfile", "-Command", command],
        input=encoded,
        text=True,
        capture_output=True,
        check=False,
        env=env,
        encoding="utf-8",
        errors="replace",
    )
    if result.returncode != 0:
        message = (result.stderr or result.stdout or "PowerShell file write failed").strip()
        raise OSError(message)


def _read_delimited(path: Path, *, delimiter: str) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file, delimiter=delimiter))

    if not rows:
        return []

    header_index = _detect_header_row(rows)
    headers = rows[header_index]
    results: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        item = {
            header: row[index] if index < len(row) else ""
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in item.values()):
            results.append(item)
    return results


def _detect_header_row(rows: list[list[str]]) -> int:
    known_headers = {
        "일별",
        "캠페인유형",
        "URL",
        "PC/모바일 매체",
        "캠페인",
        "광고그룹",
        "키워드",
        "노출수",
        "클릭수",
        "총비용",
        "총 전환수",
    }

    for index, row in enumerate(rows[:10]):
        values = {str(value).strip() for value in row if str(value).strip()}
        if len(values & known_headers) >= 2:
            return index

    for index, row in enumerate(rows[:10]):
        non_empty = [value for value in row if str(value).strip()]
        if len(non_empty) > 2:
            return index

    return 0


def _read_xlsx(path: Path, *, sheet_name: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx report files.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        return []

    headers = ["" if value is None else str(value) for value in header_row]
    results: list[dict[str, Any]] = []
    for row in rows:
        item = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if any(value not in (None, "") for value in item.values()):
            results.append(item)
    return results
